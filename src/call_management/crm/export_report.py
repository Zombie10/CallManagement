"""Build multi-sheet Excel analytics workbooks."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from call_management.crm.reports import CHANNEL_LABELS, TEMPLATE_LABELS

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(name="Arial", bold=True, color="E2E8F0", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="0F172A", size=16)
SUBTITLE_FONT = Font(name="Arial", color="64748B", size=11)
KPI_LABEL_FONT = Font(name="Arial", color="64748B", size=10)
KPI_VALUE_FONT = Font(name="Arial", bold=True, color="0891B2", size=14)
BODY_FONT = Font(name="Arial", color="1E293B", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def _fmt_duration(seconds: int | float | None) -> str:
    if seconds is None:
        return ""
    s = int(seconds)
    if s < 60:
        return f"{s}s"
    m, rem = divmod(s, 60)
    return f"{m}m {rem}s" if rem else f"{m}m"


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize_columns(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def _write_table(
    ws,
    *,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, start_row, len(headers))
    row_idx = start_row + 1
    for row in rows:
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1
    return row_idx


def _sheet_portada(
    wb: Workbook,
    *,
    tenant_name: str,
    report: dict[str, Any],
    filters: dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Reporte de análisis de llamadas"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = tenant_name
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = f"Generado: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"].font = SUBTITLE_FONT

    summary = report.get("summary") or {}
    kpis = [
        ("Total llamadas", summary.get("total_calls", 0)),
        ("Llamantes únicos", summary.get("unique_callers", 0)),
        ("Duración promedio", _fmt_duration(summary.get("avg_duration_seconds"))),
        ("Duración total", _fmt_duration(summary.get("total_duration_seconds"))),
        ("Transferencias", summary.get("handoffs", 0)),
    ]
    row = 5
    ws.cell(row=row, column=1, value="Indicadores clave").font = Font(name="Arial", bold=True, size=12)
    row += 1
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = KPI_VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Filtros aplicados").font = Font(name="Arial", bold=True, size=12)
    row += 1
    filter_lines = [
        ("Desde", filters.get("date_from") or "—"),
        ("Hasta", filters.get("date_to") or "—"),
        ("Outcomes", ", ".join(filters.get("outcomes") or []) or "Todos"),
        ("Agentes", ", ".join(filters.get("agent_instance_ids") or []) or "Todos"),
        ("Canales", ", ".join(filters.get("channels") or []) or "Todos"),
        ("Teléfono origen", filters.get("from_number") or "—"),
        ("Agrupación", report.get("group_by") or "day"),
    ]
    for label, value in filter_lines:
        ws.cell(row=row, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    channels = summary.get("channels") or []
    if channels:
        row += 1
        _write_table(
            ws,
            start_row=row,
            headers=["Canal", "Llamadas"],
            rows=[[c.get("label") or c.get("key"), c.get("count")] for c in channels],
        )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36


def _sheet_outcomes(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Outcomes")
    series = report.get("outcome_breakdown") or []
    last = _write_table(
        ws,
        start_row=1,
        headers=["Outcome", "Llamadas", "% del total", "Duración prom.", "Duración total"],
        rows=[
            [
                s.get("label"),
                s.get("count"),
                round((s.get("count", 0) / max(1, sum(x.get("count", 0) for x in series))) * 100, 1),
                _fmt_duration(s.get("avg_duration")),
                _fmt_duration(s.get("sum_duration")),
            ]
            for s in series
        ],
    )
    if len(series) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Llamadas por outcome"
        chart.y_axis.title = "Cantidad"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(series) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(series) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, f"A{last + 2}")
    _autosize_columns(ws)


def _sheet_series(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Series")
    group_by = report.get("group_by") or "day"
    series = report.get("series") or []
    _write_table(
        ws,
        start_row=1,
        headers=[f"Periodo ({group_by})", "Llamadas", "Duración prom.", "Duración total"],
        rows=[
            [
                s.get("label"),
                s.get("count"),
                _fmt_duration(s.get("avg_duration")),
                _fmt_duration(s.get("sum_duration")),
            ]
            for s in series
        ],
    )
    if len(series) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Serie por {group_by}"
        chart.y_axis.title = "Llamadas"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(series) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(series) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        ws.add_chart(chart, f"F2")
    _autosize_columns(ws)


def _sheet_pivot(wb: Workbook, report: dict[str, Any]) -> None:
    pivot = report.get("pivot")
    if not pivot:
        return
    ws = wb.create_sheet("Pivot")
    row_dim = pivot.get("row_dimension", "")
    col_dim = pivot.get("col_dimension", "")
    headers = [f"{row_dim} ↓", *pivot.get("col_labels", []), "Total"]
    rows: list[list[Any]] = []
    for ri, row_label in enumerate(pivot.get("row_labels", [])):
        cells = pivot.get("cells", [[]])[ri] if ri < len(pivot.get("cells", [])) else []
        row_total = sum(cells)
        rows.append([row_label, *cells, row_total])
    col_totals = []
    col_keys = pivot.get("col_keys") or []
    for ci in range(len(col_keys)):
        col_totals.append(sum(pivot.get("cells", [[]])[ri][ci] for ri in range(len(pivot.get("row_labels", [])))))
    rows.append(["Total", *col_totals, sum(col_totals)])
    _write_table(ws, start_row=1, headers=headers, rows=rows)
    _autosize_columns(ws, max_width=16)


def _sheet_llamadas(
    wb: Workbook,
    report: dict[str, Any],
    *,
    agent_labels: dict[str, str],
) -> None:
    ws = wb.create_sheet("Llamadas")
    detail = report.get("detail") or []
    rows = []
    for r in detail:
        agent_id = r.get("agent_instance_id") or ""
        template = r.get("transferred_to") or ""
        rows.append(
            [
                r.get("call_id"),
                r.get("start_time"),
                r.get("end_time"),
                r.get("from_number"),
                r.get("to_number"),
                CHANNEL_LABELS.get(r.get("channel") or "sip", r.get("channel")),
                (r.get("outcome") or "").replace("_", " "),
                r.get("duration_seconds"),
                _fmt_duration(r.get("duration_seconds")),
                agent_labels.get(agent_id, agent_id or "—"),
                TEMPLATE_LABELS.get(template, template or "—"),
                "Sí" if r.get("has_transcript") else "No",
                "Sí" if r.get("has_recording") else "No",
                (r.get("summary") or "")[:2000],
                (r.get("agent_notes") or "")[:1000],
            ]
        )
    _write_table(
        ws,
        start_row=1,
        headers=[
            "ID",
            "Inicio",
            "Fin",
            "Origen",
            "Destino",
            "Canal",
            "Outcome",
            "Duración (s)",
            "Duración",
            "Agente",
            "Plantilla",
            "Transcript",
            "Grabación",
            "Resumen",
            "Notas agente",
        ],
        rows=rows,
    )
    ws.freeze_panes = "A2"
    _autosize_columns(ws, max_width=40)


def build_analytics_workbook(
    *,
    tenant_name: str,
    report: dict[str, Any],
    agent_labels: dict[str, str] | None = None,
) -> bytes:
    wb = Workbook()
    labels = agent_labels or {}
    filters = report.get("filters_applied") or {}
    _sheet_portada(wb, tenant_name=tenant_name, report=report, filters=filters)
    _sheet_outcomes(wb, report)
    _sheet_series(wb, report)
    _sheet_pivot(wb, report)
    _sheet_llamadas(wb, report, agent_labels=labels)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def csv_fieldnames_spanish() -> list[tuple[str, str]]:
    return [
        ("call_id", "id_llamada"),
        ("start_time", "inicio"),
        ("end_time", "fin"),
        ("from_number", "origen"),
        ("to_number", "destino"),
        ("channel", "canal"),
        ("outcome", "outcome"),
        ("duration_seconds", "duracion_segundos"),
        ("duration_fmt", "duracion"),
        ("agent_instance_id", "agente_id"),
        ("agent_label", "agente"),
        ("transferred_to", "plantilla"),
        ("has_transcript", "transcript"),
        ("has_recording", "grabacion"),
        ("summary", "resumen"),
        ("agent_notes", "notas_agente"),
    ]